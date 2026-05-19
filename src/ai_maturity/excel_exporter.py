from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ai_maturity.taxonomy import DIMENSIONS, SUB_DIMENSIONS

# Headstorm brand colours
_NAVY  = "1A2B4C"
_ORANGE = "E75B27"
_CREAM  = "FBF8F4"
_WHITE  = "FFFFFF"

_TAB_NAMES = {
    "ai_tool_adoption":              "AI Tool Adoption",
    "prompt_context_engineering":    "Prompt + Context Eng.",
    "agent_configuration":           "Agent Configuration",
    "cicd_integration":              "CI-CD Integration",
    "ticketing_planning":            "Ticketing + Planning",
    "cross_system_connectivity":     "Cross-System Connect.",
    "quality_controls":              "Quality Controls",
    "security_compliance":           "Security + Compliance",
    "measurement_kpis":              "Measurement + KPIs",
    "ways_of_working":               "Ways of Working",
    "accountability_ownership":      "Accountability + Own.",
    "scalability_knowledge_transfer":"Scalability + KT",
}

_DIM_DISPLAY = {
    "capability":         "Capability",
    "integration":        "Integration",
    "governance":         "Governance",
    "execution_ownership":"Execution Ownership",
}

_EVIDENCE_COLS = [
    "timestamp", "session_id", "record_type", "category",
    "prompt_text", "tool_name", "agent_type", "agent_description",
    "command", "file_path", "skill_args", "raw_json",
]

_COL_WIDTHS = {
    "timestamp":        22,
    "session_id":       18,
    "record_type":      16,
    "category":         18,
    "prompt_text":      50,
    "tool_name":        18,
    "agent_type":       20,
    "agent_description":35,
    "command":          45,
    "file_path":        35,
    "skill_args":       25,
    "raw_json":         60,
}


def _derive_record_type(record: dict) -> str:
    data = record.get("data", {})
    if "prompt_text" in data:
        return "prompt"
    if "agent_type" in data:
        return "agent_spawn"
    if "subtype" in data:
        return "session_config"
    if "tool_name" in data:
        return "skill_invocation" if record.get("category") == "skill_usage" else "tool_call"
    return "unknown"


def _flatten_record(record: dict) -> dict:
    data = record.get("data", {})
    inp = data.get("input") if isinstance(data.get("input"), dict) else {}
    ts_raw = record.get("timestamp", "")
    try:
        ts = datetime.fromisoformat(ts_raw.replace("Z", "+00:00")).replace(tzinfo=None)
    except (ValueError, AttributeError):
        ts = ts_raw

    return {
        "timestamp":        ts,
        "session_id":       record.get("session_id", ""),
        "record_type":      _derive_record_type(record),
        "category":         record.get("category", ""),
        "prompt_text":      data.get("prompt_text", ""),
        "tool_name":        data.get("tool_name", ""),
        "agent_type":       data.get("agent_type", ""),
        "agent_description":data.get("agent_description", ""),
        "command":          inp.get("command", ""),
        "file_path":        inp.get("file_path", ""),
        "skill_args":       inp.get("args", ""),
        "raw_json":         json.dumps(data, ensure_ascii=False),
    }


def _header_font(bold: bool = False) -> Font:
    return Font(name="Calibri", color=_WHITE, bold=bold)


def _navy_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_NAVY)


def _orange_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_ORANGE)


def _cream_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_CREAM)


def _write_summary_tab(ws, scores: list[dict]) -> None:
    ws.title = "Summary"
    headers = ["Sub-Dimension", "Dimension", "Level", "Label", "Confidence", "Records", "Reasoning"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _header_font(bold=True)
        cell.fill = _navy_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")

    scores_by_sd = {s["sub_dimension"]: s for s in scores}
    for sd in SUB_DIMENSIONS:
        s = scores_by_sd.get(sd, {})
        dim = next((d for d, sds in DIMENSIONS.items() if sd in sds), "")
        ws.append([
            _TAB_NAMES.get(sd, sd),
            _DIM_DISPLAY.get(dim, dim),
            f"L{s.get('level', '?')}",
            s.get("level_label", ""),
            s.get("confidence", ""),
            s.get("record_count", 0),
            s.get("reasoning", ""),
        ])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 60
    ws.row_dimensions[1].height = 20


def _write_subdim_tab(ws, sd: str, score: dict | None, records: list[dict]) -> None:
    ws.title = _TAB_NAMES.get(sd, sd)
    dim = next((d for d, sds in DIMENSIONS.items() if sd in sds), "")
    s = score or {}

    summary_rows = [
        ("Sub-Dimension", _TAB_NAMES.get(sd, sd)),
        ("Dimension",     _DIM_DISPLAY.get(dim, dim)),
        ("Level",         f"L{s.get('level', '?')}"),
        ("Label",         s.get("level_label", "")),
        ("Confidence",    s.get("confidence", "")),
        ("Record Count",  s.get("record_count", len(records))),
        ("Reasoning",     s.get("reasoning", "")),
    ]
    for label, value in summary_rows:
        ws.append([label, value])
        row_idx = ws.max_row
        label_cell = ws.cell(row=row_idx, column=1)
        label_cell.font = _header_font(bold=True)
        label_cell.fill = _orange_fill()
        label_cell.alignment = Alignment(vertical="top")
        value_cell = ws.cell(row=row_idx, column=2)
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if label == "Reasoning":
            ws.row_dimensions[row_idx].height = 60

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80

    ws.append([])

    if not records:
        ws.append(["No records routed to this sub-dimension."])
        return

    header_row = ws.max_row + 1
    ws.append(_EVIDENCE_COLS)
    for i, col_name in enumerate(_EVIDENCE_COLS, start=1):
        cell = ws.cell(row=header_row, column=i)
        cell.font = _header_font(bold=True)
        cell.fill = _navy_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = _COL_WIDTHS.get(col_name, 20)

    for rec in sorted(records, key=lambda r: r.get("timestamp", "")):
        flat = _flatten_record(rec)
        ws.append([flat[col] for col in _EVIDENCE_COLS])
        data_row = ws.max_row
        if (data_row - header_row) % 2 == 1:
            for col in range(1, len(_EVIDENCE_COLS) + 1):
                ws.cell(row=data_row, column=col).fill = _cream_fill()
        for col_idx, col_name in enumerate(_EVIDENCE_COLS, start=1):
            if col_name in ("prompt_text", "raw_json", "command"):
                ws.cell(row=data_row, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[data_row].height = 40

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def export_excel(
    records: list[dict],
    scores: list[dict],
    developer: dict,
    output_path: Path,
) -> None:
    if not scores:
        raise ValueError("No scores found — run 'assess' before export-excel")

    scores_by_sd = {s["sub_dimension"]: s for s in scores}
    records_by_sd: dict[str, list[dict]] = {sd: [] for sd in SUB_DIMENSIONS}
    for rec in records:
        sd = rec.get("sub_dimension", "")
        if sd in records_by_sd:
            records_by_sd[sd].append(rec)

    wb = openpyxl.Workbook()
    summary_ws = wb.active
    _write_summary_tab(summary_ws, scores)

    for sd in SUB_DIMENSIONS:
        ws = wb.create_sheet()
        _write_subdim_tab(ws, sd, scores_by_sd.get(sd), records_by_sd[sd])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
