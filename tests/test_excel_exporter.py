from __future__ import annotations

from pathlib import Path
import pytest
import openpyxl
from ai_maturity.excel_exporter import export_excel, _derive_record_type, _flatten_record

# ── fixtures ────────────────────────────────────────────────

def _make_record(category, sub_dimension, data, session_id="s1",
                 timestamp="2026-05-01T09:00:00.000Z"):
    return {
        "id": f"{category}-{sub_dimension}",
        "category": category,
        "sub_dimension": sub_dimension,
        "dimension": "capability",
        "team": "eng",
        "user": "alice",
        "session_id": session_id,
        "timestamp": timestamp,
        "source": "session",
        "data": data,
        "metadata": {"cwd": "/home", "version": "1"},
    }

PROMPT_REC = _make_record("prompts", "ai_tool_adoption",
                           {"prompt_text": "help me debug"})
TOOL_REC   = _make_record("tool_usage", "cicd_integration",
                           {"tool_name": "Bash", "input": {"command": "pytest tests/"}})
AGENT_REC  = _make_record("agent_delegation", "agent_configuration",
                           {"tool_name": "Agent", "agent_type": "general-purpose",
                            "agent_description": "run tests", "agent_prompt_summary": "...",
                            "parallel_agents": None, "input": {}})
SKILL_REC  = _make_record("skill_usage", "agent_configuration",
                           {"tool_name": "Skill", "input": {"skill": "tdd", "args": "write tests"}})
SESSION_REC = _make_record("tool_usage", "ways_of_working",
                            {"subtype": "hook", "hook_count": 2, "hooks": [], "content": ""})

ALL_SCORES = [
    {"sub_dimension": sd, "dimension": "capability", "level": 2,
     "level_label": "Integrated", "confidence": "medium",
     "reasoning": f"Reasoning for {sd}.", "evidence": [], "matched_signals": [],
     "record_count": 3}
    for sd in [
        "ai_tool_adoption", "prompt_context_engineering", "agent_configuration",
        "cicd_integration", "ticketing_planning", "cross_system_connectivity",
        "quality_controls", "security_compliance", "measurement_kpis",
        "ways_of_working", "accountability_ownership", "scalability_knowledge_transfer",
    ]
]

DEV = {"name": "alice", "email": "alice@co.com", "team": "eng"}

# ── unit tests ───────────────────────────────────────────────

def test_derive_record_type_prompt():
    assert _derive_record_type(PROMPT_REC) == "prompt"

def test_derive_record_type_tool_call():
    assert _derive_record_type(TOOL_REC) == "tool_call"

def test_derive_record_type_agent_spawn():
    assert _derive_record_type(AGENT_REC) == "agent_spawn"

def test_derive_record_type_skill_invocation():
    assert _derive_record_type(SKILL_REC) == "skill_invocation"

def test_derive_record_type_session_config():
    assert _derive_record_type(SESSION_REC) == "session_config"

def test_flatten_prompt():
    flat = _flatten_record(PROMPT_REC)
    assert flat["prompt_text"] == "help me debug"
    assert flat["tool_name"] == ""
    assert flat["command"] == ""
    assert flat["record_type"] == "prompt"
    assert "raw_json" in flat

def test_flatten_tool_call():
    flat = _flatten_record(TOOL_REC)
    assert flat["tool_name"] == "Bash"
    assert flat["command"] == "pytest tests/"
    assert flat["prompt_text"] == ""
    assert flat["record_type"] == "tool_call"

def test_flatten_agent_spawn():
    flat = _flatten_record(AGENT_REC)
    assert flat["agent_type"] == "general-purpose"
    assert flat["agent_description"] == "run tests"
    assert flat["record_type"] == "agent_spawn"

def test_flatten_skill_invocation():
    flat = _flatten_record(SKILL_REC)
    assert flat["skill_args"] == "write tests"
    assert flat["record_type"] == "skill_invocation"

# ── integration tests ────────────────────────────────────────

def test_export_excel_creates_file(tmp_path):
    out = tmp_path / "alice_report.xlsx"
    records = [PROMPT_REC, TOOL_REC, AGENT_REC, SKILL_REC]
    export_excel(records, ALL_SCORES, DEV, out)
    assert out.exists()

def test_export_excel_tab_count(tmp_path):
    out = tmp_path / "alice_report.xlsx"
    export_excel([PROMPT_REC], ALL_SCORES, DEV, out)
    wb = openpyxl.load_workbook(out)
    # 1 Summary + 12 sub-dimension tabs
    assert len(wb.sheetnames) == 13

def test_export_excel_summary_tab_first(tmp_path):
    out = tmp_path / "alice_report.xlsx"
    export_excel([PROMPT_REC], ALL_SCORES, DEV, out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames[0] == "Summary"

def test_export_excel_all_subdim_tabs_present(tmp_path):
    out = tmp_path / "alice_report.xlsx"
    export_excel([], ALL_SCORES, DEV, out)
    wb = openpyxl.load_workbook(out)
    expected_subdim_tabs = [
        "AI Tool Adoption", "Prompt + Context Eng.", "Agent Configuration",
        "CI-CD Integration", "Ticketing + Planning", "Cross-System Connect.",
        "Quality Controls", "Security + Compliance", "Measurement + KPIs",
        "Ways of Working", "Accountability + Own.", "Scalability + KT",
    ]
    assert wb.sheetnames[0] == "Summary"
    assert wb.sheetnames[1:] == expected_subdim_tabs

def test_export_excel_raises_without_scores(tmp_path):
    out = tmp_path / "alice_report.xlsx"
    with pytest.raises(ValueError, match="scores"):
        export_excel([PROMPT_REC], [], DEV, out)

def test_export_excel_summary_tab_content(tmp_path):
    out = tmp_path / "alice_report.xlsx"
    export_excel([PROMPT_REC], ALL_SCORES, DEV, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Summary"]
    assert ws.cell(row=1, column=1).value == "Sub-Dimension"
    assert ws.cell(row=2, column=1).value == "AI Tool Adoption"
    assert ws.cell(row=2, column=3).value == "L2"
